import nltk
nltk.download('stopwords')
from nltk.corpus import stopwords
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
import string
# from sklearn.metrics.pairwise import cosine_similarity
from sklearn.feature_extraction.text import CountVectorizer
stopwords = stopwords.words('english')
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from nltk.translate import bleu
from nltk.translate.bleu_score import SmoothingFunction


def cosineSimilarity(base, response):
    sentences = [base, response]
    def clean_string(text):
        text = ''.join([word for word in text if word not in string.punctuation])
        text = text.lower()
        text = ' '.join([word for word in text.split() if word not in stopwords])
        return text
    cleaned = list(map(clean_string, sentences))
    vectorizer = CountVectorizer().fit_transform(cleaned)
    vectors = vectorizer.toarray()
    csim = cosine_similarity(vectors)
    def cosine_sim_vectors(vec1, vec2):
        vec1 = vec1.reshape(1, -1)
        vec2 = vec2.reshape(1, -1)
        return cosine_similarity(vec1, vec2)[0][0]
    return (cosine_sim_vectors(vectors[0], vectors[1]))
def bertMethod1(base,response):
    sen = [
        base,response
    ]
    model = SentenceTransformer('bert-base-nli-mean-tokens')
    sen_embeddings = model.encode(sen)
    return(cosine_similarity(
        [sen_embeddings[0]],
        sen_embeddings[1:]
    ))
def newMethod1(base,response):
    sen = [
        base,response
    ]
    model = SentenceTransformer('all-mpnet-base-v2')
    sen_embeddings = model.encode(sen)
    return(cosine_similarity(
        [sen_embeddings[0]],
        sen_embeddings[1:]
    ))
def newMethod2(base,response):
    sen = [
        base,response
    ]
    model = SentenceTransformer('all-distilroberta-v1')
    sen_embeddings = model.encode(sen)
    return(cosine_similarity(
        [sen_embeddings[0]],
        sen_embeddings[1:]
    ))
def newMethod3(base,response):
    
    sen = [
        base,response
    ]
    model = SentenceTransformer('all-MiniLM-L12-v2')
    sen_embeddings = model.encode(sen)
    return(cosine_similarity(
        [sen_embeddings[0]],
        sen_embeddings[1:]
    ))
def newMethod4(base,response):
    sen = [
        base,response
    ]
    model = SentenceTransformer('all-MiniLM-L6-v2')
    sen_embeddings = model.encode(sen)
    return(cosine_similarity(
        [sen_embeddings[0]],
        sen_embeddings[1:]
    ))
def integrated_value_of_bert_and_newMethod1(B,N,numberOfBaseWords,numberOfResponseWords):
    difference=numberOfBaseWords-numberOfResponseWords
    if(abs(difference)>=(0.55*numberOfBaseWords)):
        return 0.6*min(B,N)+0.4*max(B,N)
    else:
        return 0.6*N+0.4*B

def show_result():
    total_no_of_testCases = 40
    list_of_integrated_results=[]
    list_of_bert_results=[]
    list_of_newMethod1_results=[]
    list_of_noOfBaseWords=[]
    list_of_noOfResponseWords=[]
    for i in range(1,total_no_of_testCases+1):
        test_case="Test Number-"+str(i)
        b="base"+str(i)+".txt"
        r="trail_"+str(i)+".txt"
        with open(b, encoding = "UTF-8") as file:
            base = file.read()
        with open(r, encoding = "UTF-8") as file:
            response = file.read()
        cosinesim_result=(cosineSimilarity(base,response))*100
        bertmethod1_result=(bertMethod1(base,response)[0][0])*100
        newmethod1_result = (newMethod1(base,response)[0][0]) * 100
#         newmethod2_result = (newMethod2(base, response)[0][0]) * 100
#         newmethod3_result = (newMethod3(base, response)[0][0]) * 100
        # newmethod4_result = (test.newMethod4(base, response)[0][0]) * 100
        def BaseWordCount(base):
            list_base = base.split(" ")
            return (len(list_base))
        no_of_base_words=BaseWordCount(base)
        def ResponseWordCount(response):
            list_response = response.split(" ")
            return (len(list_response))
        no_of_response_words=ResponseWordCount(response)
        
        integrated_result= integrated_value_of_bert_and_newMethod1(bertmethod1_result,newmethod1_result,no_of_base_words,no_of_response_words)
        outfile=open('allData.txt', 'a')
        outfile.write(test_case)
        outfile.write("\nNumber Of Base Words: "+str(no_of_base_words)+"\n")
        outfile.write("Number Of Response Words: "+str(no_of_response_words)+"\n")
        outfile.write("Base: "+base+"\n")
        outfile.write("Response: "+response+"\n")
        outfile.write("Cosine Similarity Result: "+str(cosinesim_result)+"\n")
        outfile.write("BERT Similarity Result: "+str(bertmethod1_result)+"\n")
        outfile.write("newMethod1 Similarity Result: " + str(newmethod1_result) + "\n")
        outfile.write("INTEGRATED RESULT: "+str(integrated_result)+"\n\n\n\n\n\n")
        
        
        
        list_of_integrated_results.append(integrated_result)
        list_of_bert_results.append(bertmethod1_result)
        list_of_newMethod1_results.append(newmethod1_result)
        list_of_noOfBaseWords.append(no_of_base_words)
        list_of_noOfResponseWords.append(no_of_response_words)
        
        
             
        
        
#         outfile.write("newMethod2 Similarity Result: " + str(newmethod2_result) + "\n")
#         outfile.write("newMethod3 Similarity Result: " + str(newmethod3_result) + "\n\n\n")
        # outfile.write("newMethod4 Similarity Result: " + str(newmethod4_result) + "\n\n\n")
        outfile.close()
    book=openpyxl.load_workbook('allData.xlsx')
    sheet=book['Sheet2']
    sheet.delete_rows(1,sheet.max_row)
    data=pd.DataFrame({'Integrated Result':list_of_integrated_results,'Bert Result':list_of_bert_results,'NewMethod Result':list_of_newMethod1_results,'No of Base Words':list_of_noOfBaseWords,'No of Response words':list_of_noOfResponseWords})
    datatoexcel=pd.ExcelWriter("allData.xlsx",engine='xlsxwriter')
    data.to_excel(datatoexcel,sheet_name='Sheet2')
    datatoexcel.save()
    
    
        
        
def delete_result():
    outfile = open(r'C:\Users\rajva\OneDrive\Desktop\NLP MODEL\allData.txt', 'w')
    outfile.write("")
    outfile.write("")
    outfile.close()
delete_result()
